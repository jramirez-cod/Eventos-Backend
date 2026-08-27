from io import BytesIO

import pytest
from openpyxl import load_workbook

from test.modules.eventos.conftest import (
    crear_evento_http,
    future_date,
    seed_event_actor,
)


pytestmark = pytest.mark.asyncio


async def _seed_eventos(client, session_factory):
    async with session_factory() as session:
        _, headers = await seed_event_actor(session)
    first = await crear_evento_http(
        client,
        headers,
        session_factory,
        nombre_evento="Congreso Comercial",
        politica_fecha_inicio=future_date(5),
        politica_fecha_fin=future_date(7),
    )
    second = await crear_evento_http(
        client,
        headers,
        session_factory,
        nombre_evento="Foro Virtual",
        politica_fecha_inicio=future_date(10),
        politica_fecha_fin=future_date(12),
    )
    await client.patch(
        f"/api/v1/eventos/{second['id_evento']}/finalizar",
        headers=headers,
        json={},
    )
    return headers, first, second


async def test_listar_buscar_y_paginar(client, session_factory) -> None:
    headers, first, _ = await _seed_eventos(client, session_factory)
    search = await client.get(
        "/api/v1/eventos", headers=headers, params={"search": "Comercial"}
    )
    assert search.status_code == 200
    assert search.json()["total"] == 1
    assert search.json()["items"][0]["id_evento"] == first["id_evento"]

    page = await client.get(
        "/api/v1/eventos",
        headers=headers,
        params={"page": 2, "page_size": 1},
    )
    assert page.json()["total"] == 2
    assert page.json()["pages"] == 2
    assert len(page.json()["items"]) == 1


async def test_filtro_estado(client, session_factory) -> None:
    headers, _, second = await _seed_eventos(client, session_factory)
    response = await client.get(
        "/api/v1/eventos",
        headers=headers,
        params={"estado": "FINALIZADO"},
    )
    assert response.status_code == 200
    assert response.json()["total"] == 1
    assert response.json()["items"][0]["id_evento"] == second["id_evento"]


async def test_filtro_area(client, session_factory) -> None:
    headers, first, second = await _seed_eventos(client, session_factory)
    response = await client.get(
        "/api/v1/eventos",
        headers=headers,
        params={"id_area": first["id_area"]},
    )
    assert response.status_code == 200
    ids = {item["id_evento"] for item in response.json()["items"]}
    assert first["id_evento"] in ids
    assert second["id_evento"] not in ids


async def test_filtro_fecha_utiliza_solapamiento(client, session_factory) -> None:
    headers, first, _ = await _seed_eventos(client, session_factory)
    response = await client.get(
        "/api/v1/eventos",
        headers=headers,
        params={
            "fecha_desde": future_date(7),
            "fecha_hasta": future_date(9),
        },
    )
    assert response.status_code == 200
    assert response.json()["total"] == 1
    assert response.json()["items"][0]["id_evento"] == first["id_evento"]


async def test_filtro_fecha_invertido_recibe_400(client, session_factory) -> None:
    headers, _, _ = await _seed_eventos(client, session_factory)
    response = await client.get(
        "/api/v1/eventos",
        headers=headers,
        params={
            "fecha_desde": future_date(10),
            "fecha_hasta": future_date(5),
        },
    )
    assert response.status_code == 400


async def test_exportar_xlsx_respeta_filtros_y_columnas(
    client, session_factory
) -> None:
    headers, first, _ = await _seed_eventos(client, session_factory)
    response = await client.get(
        "/api/v1/eventos/exportar",
        headers=headers,
        params={"search": "Comercial"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    sheet = workbook["Eventos"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "ID",
        "Nombre",
        "Descripción",
        "Área",
        "Política: Fecha inicio",
        "Política: Fecha fin",
        "Estado",
    )
    assert len(rows) == 2
    assert rows[1][0] == first["id_evento"]
    assert rows[1][1] == "Congreso Comercial"
    workbook.close()
